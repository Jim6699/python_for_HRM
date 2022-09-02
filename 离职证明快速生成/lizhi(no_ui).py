from docx import Document
from openpyxl import load_workbook
import os
import datetime
import time
# 结合路径判断生成文件夹，规避程序报错而终止的风险
if not os.path.exists('./' + '离职证明导出'):
    os.mkdir('./' + '离职证明导出')
    print("创建目录成功")
import tkinter as tk
from tkinter import filedialog
print("请选择花名册文件")
try:
    root = tk.Tk()
    root.withdraw()
    Filepath = filedialog.askopenfilename()
    workbook = load_workbook(Filepath)
except:
    print("未选择文件，请重新选择！")
    root = tk.Tk()
    root.withdraw()
    Filepath = filedialog.askopenfilename()
    workbook = load_workbook(Filepath)
sheet = workbook["离职"]
numbers = int(sheet.max_row)-1
who = input("请输入姓名：")
summit = 0
for table_row in range(2, sheet.max_row + 1):
    name = str(sheet.cell(row=table_row, column=6).value)
    if name == who:
        print("找到1个符合条件")
        try:
            try:
                print("=====入职日期=====")
                date_kaishi =  sheet['I{}'.format(table_row)].value
                try:
                    date_kaishi_year = "20"+str(date_kaishi.strftime('%y')) 
                    date_kaishi_month = str(date_kaishi.strftime('%m'))
                    date_kaishi_day = str(date_kaishi.strftime('%d'))
                except:
                    date_kaishi_year = datetime.datetime.strptime(date_kaishi,'%Y-%m-%d').year
                    date_kaishi_month = datetime.datetime.strptime(date_kaishi,'%Y-%m-%d').month
                    date_kaishi_day = datetime.datetime.strptime(date_kaishi,'%Y-%m-%d').day
                print(date_kaishi_year)
                print(date_kaishi_month)
                print(date_kaishi_day)
                print("=====离职日期=====")
                date_zhongzhi = sheet['J{}'.format(table_row)].value
                try:
                    date_zhongzhi_year = "20"+str(date_zhongzhi.strftime('%y')) 
                    date_zhongzhi_month = str(date_zhongzhi.strftime('%m'))
                    date_zhongzhi_day = str(date_zhongzhi.strftime('%d'))
                except:
                    date_zhongzhi_year = datetime.datetime.strptime(date_zhongzhi,'%Y-%m-%d').year
                    date_zhongzhi_month = datetime.datetime.strptime(date_zhongzhi,'%Y-%m-%d').month
                    date_zhongzhi_day = datetime.datetime.strptime(date_zhongzhi,'%Y-%m-%d').day
                print(date_zhongzhi_year)
                print(date_zhongzhi_month)
                print(date_zhongzhi_day)
                print("=====身份证=====")
                date_ID = sheet['H{}'.format(table_row)].value
                print(date_ID)
                print("=====岗位=====")
                GW = sheet['E{}'.format(table_row)].value
                print(GW)
                wordfile = Document('./' + '离职证明【模板勿动】.docx')
                all_paragraphs = wordfile.paragraphs
                for paragraph in all_paragraphs:
                    for run in paragraph.runs:
                        try:
                            if "奰" in run.text:
                                run.text = run.text.replace("奰", date_kaishi_year)
                            elif "躄" in run.text:  
                                run.text = run.text.replace("躄", date_kaishi_month)
                            elif "罍" in run.text:  
                                run.text = run.text.replace("罍", date_kaishi_day)
                            elif "颣" in run.text:  
                                run.text = run.text.replace("颣", date_zhongzhi_year)
                            elif "薐" in run.text:   
                                run.text = run.text.replace("薐", date_zhongzhi_month)
                            elif "豳" in run.text:
                                run.text = run.text.replace("豳", date_zhongzhi_day)
                            elif "懿" in run.text:
                                run.text = run.text.replace("懿", str(date_ID))
                            elif "鰘" in run.text:
                                run.text = run.text.replace("鰘", str(GW))    
                            elif "礥" in run.text:
                                run.text = run.text.replace("礥", str(who))    
                        except Exception as e:
                            print("替换文本出错："+str(e)) 
                wordfile.save('./' + f'离职证明导出/{table_row}_{name}_离职证明.docx')
                print(f"{table_row}_{name}_离职证明.docx | 另存成功")
                summit += 1
            except Exception as e:
                print("内出错："+str(e)) 
        except Exception as e:
            print("外出错："+str(e)) 
if summit == 0:
    print(f"未查到 {who}！！！！！\n")
elif summit == 1:
    print("导出一份，结束！\n")
elif summit >= 2:
    print("结束！\n")
    print(f"本次共导出 {summit} 个 {who}的离职文件,请注意筛选！！！！！\n")
input ("Please Enter to close this exe:")
