# -*- coding: utf-8 -*-
import wx
import time
import tkinter as tk
from tkinter import filedialog

from docx import Document
from openpyxl import load_workbook
import os
import datetime

def lzzm(Filepath,who):
    if not os.path.exists('./' + '离职证明导出'):
        os.mkdir('./' + '离职证明导出')
        contents.AppendText("创建目录成功\n")
    workbook = load_workbook(Filepath)
    sheet = workbook["离职"]
    summit = 0
    for table_row in range(2, sheet.max_row + 1):
        name = str(sheet.cell(row=table_row, column=6).value)
        if name == who:
            contents.AppendText("查到一个符合条件\n")
            try:
                try:
                    contents.AppendText("=====入职日期=====\n")
                    date_kaishi =  sheet['I{}'.format(table_row)].value
                    try:
                        date_kaishi_year = "20"+str(date_kaishi.strftime('%y')) 
                        date_kaishi_month = str(date_kaishi.strftime('%m'))
                        date_kaishi_day = str(date_kaishi.strftime('%d'))
                    except:
                        date_kaishi_year = datetime.datetime.strptime(date_kaishi,'%Y-%m-%d').year
                        date_kaishi_month = datetime.datetime.strptime(date_kaishi,'%Y-%m-%d').month
                        date_kaishi_day = datetime.datetime.strptime(date_kaishi,'%Y-%m-%d').day
                    contents.AppendText(date_kaishi_year)
                    contents.AppendText(date_kaishi_month)
                    contents.AppendText(date_kaishi_day+"\n")
                    contents.AppendText("=====离职日期=====\n")
                    date_zhongzhi = sheet['J{}'.format(table_row)].value
                    try:
                        date_zhongzhi_year = "20"+str(date_zhongzhi.strftime('%y')) 
                        date_zhongzhi_month = str(date_zhongzhi.strftime('%m'))
                        date_zhongzhi_day = str(date_zhongzhi.strftime('%d'))
                    except:
                        date_zhongzhi_year = datetime.datetime.strptime(date_zhongzhi,'%Y-%m-%d').year
                        date_zhongzhi_month = datetime.datetime.strptime(date_zhongzhi,'%Y-%m-%d').month
                        date_zhongzhi_day = datetime.datetime.strptime(date_zhongzhi,'%Y-%m-%d').day
                    contents.AppendText(date_zhongzhi_year)
                    contents.AppendText(date_zhongzhi_month)
                    contents.AppendText(date_zhongzhi_day+"\n")
                    contents.AppendText("=====身份证=====\n")   
                    date_ID = sheet['H{}'.format(table_row)].value
                    contents.AppendText(str(date_ID)+"\n")
                    contents.AppendText("=====岗位=====\n") 
                    GW = sheet['E{}'.format(table_row)].value
                    contents.AppendText(str(GW)+"\n")
                    contents.AppendText("即将生成文件...\n") 
                    wordfile = Document('./' + '离职证明【模板勿动】.docx')
                    all_paragraphs = wordfile.paragraphs
                    for paragraph in all_paragraphs:
                        for run in paragraph.runs:
                            try:
                                if "奰" in run.text:
                                    run.text = run.text.replace("奰", date_kaishi_year)
                                elif "躄" in run.text:  
                                    run.text = run.text.replace("躄", date_kaishi_month)
                                elif "罍" in run.text:  
                                    run.text = run.text.replace("罍", date_kaishi_day)
                                elif "颣" in run.text:  
                                    run.text = run.text.replace("颣", date_zhongzhi_year)
                                elif "薐" in run.text:   
                                    run.text = run.text.replace("薐", date_zhongzhi_month)
                                elif "豳" in run.text:
                                    run.text = run.text.replace("豳", date_zhongzhi_day)
                                elif "懿" in run.text:
                                    run.text = run.text.replace("懿", str(date_ID))
                                elif "鰘" in run.text:
                                    run.text = run.text.replace("鰘", str(GW))    
                                elif "礥" in run.text:
                                    run.text = run.text.replace("礥", str(who))    
                            except Exception as e:
                                print("替换文本出错："+str(e)) 
                    wordfile.save('./' + f'离职证明导出/{table_row}_{name}_离职证明.docx')
                    contents.AppendText(f"{table_row}_{name}_离职证明.docx | 另存成功\n")
                    summit += 1
                except Exception as e:
                    contents.AppendText("内出错："+str(e)+"\n") 
            except Exception as e:
                contents.AppendText("外出错："+str(e)+"\n") 
    if summit == 0:
        contents.AppendText(f"未查到 {who}！！！！！\n")
    elif summit == 1:
        contents.AppendText("导出一份，结束！\n")
    elif summit >= 2:
        contents.AppendText("结束！\n")
        contents.AppendText(f"本次共导出 {summit} 个 {who}的离职文件,请注意筛选！！！！！\n")

        
def choice(event):
    print("请选择文件")
    root = tk.Tk()
    root.withdraw()
    Filepath = filedialog.askopenfilename() #获得选择好的文件
    print('Filepath:',Filepath)
    filename.SetValue(Filepath)
    
def Go(event):
    if filename.GetValue() == "" or filename.GetValue() == "未选择文件":
        contents.AppendText("未选择文件，请选择后运行\n")
    elif filename1.GetValue() == "请输入姓名" or filename1.GetValue() == "":
        contents.AppendText("未输入姓名，请输入后运行\n")
    else:
        try:
            lzzm(Filepath=filename.GetValue(),who=filename1.GetValue())
        except Exception as e:
            contents.AppendText(f"{e}：运行失败！\n")

  
app = wx.App()
win = wx.Frame(None,title = "离职证明快速生成小助手", size=(535,520))
bkg = wx.Panel(win)
#设置icon
##icon = wx.Icon(r'logo.ico')
##win.SetIcon(icon)
#设置透明度
win.SetTransparent(230)
loadButton = wx.Button(bkg, label = '选择文件')
loadButton.Bind(wx.EVT_BUTTON,choice)
saveButton = wx.Button(bkg, label = '开始运行')
saveButton.Bind(wx.EVT_BUTTON,Go)
filename = wx.TextCtrl(bkg,value = "未选择文件", style = wx.TE_READONLY)
filename1 = wx.TextCtrl(bkg,value = "请输入姓名")
contents = wx.TextCtrl(bkg,value = "程序指南：\n1.选择文件选择花名册后\n2.请输入需要制作离职证明的姓名\n3.点击开始运行，如果在离职的sheet里面存在此人会导出，反之不存在！\n ====================\n", style = wx.TE_MULTILINE | wx.HSCROLL | wx.TE_READONLY)
hbox = wx.BoxSizer()
hbox.Add(filename, proportion =1, flag = wx.EXPAND)
hbox.Add(loadButton, proportion =0,flag = wx.LEFT, border = 5)
pbox = wx.BoxSizer()
pbox.Add(filename1, proportion =1, flag = wx.EXPAND)
pbox.Add(saveButton, proportion =0,flag = wx.LEFT, border = 5)
vbox = wx.BoxSizer(wx.VERTICAL)
vbox.Add(hbox,proportion = 0,flag = wx.EXPAND | wx.ALL, border = 5)
vbox.Add(pbox,proportion = 0,flag = wx.EXPAND | wx.ALL, border = 5)
vbox.Add(contents, proportion = 1,flag=wx.EXPAND | wx.LEFT | wx.BOTTOM | wx.RIGHT, border = 5)
bkg.SetSizer(vbox)
win.Show()
app.MainLoop()
